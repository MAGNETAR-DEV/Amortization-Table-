import openpyxl as oe
import random
import time
import datetime

def str_time_prop(start, end, format, prop):
	stime = time.mktime(time.strptime(start, format))
	etime = time.mktime(time.strptime(end, format))
	ptime = stime + prop * (etime - stime)
	return time.strftime(format, time.localtime(ptime))


def random_date(start, end, prop):
	return str_time_prop(start, end, '%Y-%m-%d %I:%M %p', prop)



while True: 
	number_of_set = [int(i) for i in input("How many attribute sets do you want to define (A, B, C etc) (Like this 1 2 3)?") if i.isnumeric()]

	starting_date = random_date("2008-01-01 1:30 PM", "2009-01-01 4:50 AM", random.random())
	for n in number_of_set: 
		for i in range(n): 
			# accep the input (s)
			principle_value = round(float(random.choice([i*1.99 for i in range(1000, 1000000)])))
			interest_rate = round(float(random.choice([i*1.99 for i in range(1, 3)])))
			interest_amount = principle_value * (interest_rate / 100)
			number_of_periods = random.choice(range(10,100))

			amount_of_installment = 0
			while amount_of_installment < interest_amount and amount_of_installment < principle_value: 
				amount_of_installment = (interest_amount  * 3) / 2

			# display the line for formatting table
			print("|{:^100}|".format("-----------------------------------------------------------------------------------------------------------------------------------"))
			# display column headers
			print("|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}".format("Date", "Period Number", "Amortization", "Interest Amount", "Redemption Amount", "Balance Remaining"))

			# open a new workbook 
			excel_workbook = oe.Workbook()
			# start with the current sheet first sheet
			excel_sheet = excel_workbook.active

			# indexes for accessing rows and columns start from 1, So first cell is at location (1,1)
			# write the column headers in excel file
			excel_sheet.cell(row=1,column=1).value = "Starting Date"

			excel_sheet.cell(row=1,column=2).value = "Periord Number"

			excel_sheet.cell(row=1,column=3).value = "Amortization"

			excel_sheet.cell(row=1,column=4).value = "Interest Amount"

			excel_sheet.cell(row=1,column=5).value = "Redemption Amount"

			excel_sheet.cell(row=1, column=6).value = "Balance Remaining"



			# set width of columns
			excel_sheet.column_dimensions['A'].width = float(20)
			excel_sheet.column_dimensions['B'].width = float(20)
			excel_sheet.column_dimensions['C'].width = float(20)
			excel_sheet.column_dimensions['D'].width = float(20)
			excel_sheet.column_dimensions['E'].width = float(20)
			excel_sheet.column_dimensions['F'].width = float(20)

			# counter variables
			row_counter = 2 
			period_counter = 0

			date = starting_date
			# loop to construct table
			# while number of periods are not complete
			while principle_value > 0:
				# compute the interest amount
				interest_amount = principle_value * (interest_rate / 100)
				# compute the amount that will actually get reduced from loan amount
				redemption_value = amount_of_installment - interest_amount 
				# compute the remaining loan amount
				principle_value = principle_value - redemption_value
				# increment the period counter 
				period_counter += 1

				# save data to row(s), round all the values to two decimal place with round() function
				excel_sheet.cell(row=row_counter, column=1).value = date.split(" ")[0]
				excel_sheet.cell(row=row_counter, column=2).value = round(period_counter, 2)
				excel_sheet.cell(row=row_counter, column=3).value = round(amount_of_installment, 2)
				excel_sheet.cell(row=row_counter, column=4).value = round(interest_amount, 2)
				excel_sheet.cell(row=row_counter, column=5).value = round(redemption_value, 2)
				excel_sheet.cell(row=row_counter, column=6).value = round(principle_value, 2)

				year = int(date.split()[0].split("-")[0])
				month = int(date.split()[0].split("-")[1])
				day = int(date.split()[0].split("-")[2])
				hour = int(date.split()[1].split(":")[0])
				minute = int(date.split()[1].split(":")[1]);
				date = datetime.datetime(year, month, day, hour,minute) 
				date += datetime.timedelta(days=1) 
				date = str(date)
				# increment the row counter to store next data record on next row
				row_counter += 1
				

				print("|{:^100}|".format("-----------------------------------------------------------------------------------------------------------------------------------"))
				print("|{:^20}|{:^20}|{:^20.2f}|{:^20.2f}|{:^20.2f}|{:^20.2f}|".format(str(date).split(" ")[0],str(period_counter), amount_of_installment, interest_amount, redemption_value, principle_value))

			print("|{:^100}|".format("-----------------------------------------------------------------------------------------------------------------------------------"))



			# save the file
			excel_workbook.save("at" + str(i+1) + str(number_of_set.index(n)) + ".xlsx")
